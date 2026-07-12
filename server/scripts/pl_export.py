# -*- coding: utf-8 -*-
# รายงานสรุปรายรับ–รายจ่าย (P&L) รายเดือน — จัดรูปแบบสวยงาม
# Usage: python pl_export.py <data.json> <out.xlsx>
import sys, json, warnings
warnings.simplefilter("ignore")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

dataf, out = sys.argv[1], sys.argv[2]
d = json.load(open(dataf, encoding="utf-8-sig"))

TH = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
      "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
y, m = d["month"].split("-")
month_th = f"{TH[int(m)]} {int(y) + 543}"

FONT = "Tahoma"
NAVY = "1E3A5F"; GREEN = "0B7A3B"; RED = "B42318"; GREY = "6B7280"
MONEY = '#,##0.00;[Red](#,##0.00)'
thin = Side(style="thin", color="D8DEE9")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "สรุปรายรับรายจ่าย"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 3

def cell(coord, val, *, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws[coord]; c.value = val
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = border
    return c

R = Alignment(horizontal="right", vertical="center")
L = Alignment(horizontal="left", vertical="center")
C = Alignment(horizontal="center", vertical="center")

# ── หัวรายงาน ──
ws.merge_cells("B2:D2")
cell("B2", d.get("org_name", ""), font=Font(name=FONT, size=13, bold=True, color=NAVY), align=C)
ws.merge_cells("B3:D3")
cell("B3", f"รายงานสรุปรายรับ–รายจ่าย  ประจำเดือน {month_th}", font=Font(name=FONT, size=11, color=GREY), align=C)
ws.row_dimensions[2].height = 22

row = 5
def section(title):
    global row
    ws.merge_cells(f"B{row}:D{row}")
    cell(f"B{row}", title, font=Font(name=FONT, size=10.5, bold=True, color="FFFFFF"), fill=NAVY, align=L)
    ws.row_dimensions[row].height = 20
    row += 1

def line(label, amount, *, bold=False, color=None, indent=False, sign=""):
    global row
    cell(f"B{row}", ("      " if indent else "") + label,
         font=Font(name=FONT, size=10.5, bold=bold, color=color or "111827"), align=L, border=box)
    cell(f"C{row}", None, border=box)
    disp = amount if sign != "-" else -abs(amount)
    cell(f"D{row}", disp, font=Font(name=FONT, size=10.5, bold=bold, color=color or "111827"),
         align=R, fmt=MONEY, border=box)
    ws.row_dimensions[row].height = 18
    row += 1

def detail(label, amount):
    global row
    cell(f"B{row}", "      • " + (label or "(ไม่ระบุ)"), font=Font(name=FONT, size=9.5, color=GREY), align=L, border=box)
    cell(f"C{row}", amount, font=Font(name=FONT, size=9.5, color=GREY), align=R, fmt=MONEY, border=box)
    cell(f"D{row}", None, border=box)
    ws.row_dimensions[row].height = 16
    row += 1

# ── รายรับ ──
section("รายรับ")
line(f"รายรับจากโรงงาน (Amphenol)", d["revenue"], bold=True, color=GREEN)

# ── หักต้นทุน/ค่าใช้จ่าย ──
section("หัก ต้นทุนและค่าใช้จ่าย")
line("ค่าแรงสมาชิก (ค่าตัด)", d["wage"], color=RED, sign="-")
line(f"ภาษี ณ ที่จ่าย {d.get('tax_pct', 3)}%", d["tax"], color=RED, sign="-")

# ค่าตอบแทนผู้บริหาร (ฐาน + จ่ายให้สมาชิก/ผู้บริหาร)
line("ค่าตอบแทนผู้บริหาร (รวม)", d["manager_comp"], color=RED, sign="-")
for mg in d.get("manager_lines", []):
    if mg["computed"]:
        detail(f'{mg["name"]}{(" · " + mg["role"]) if mg.get("role") else ""}', mg["computed"])
for e in d.get("comp_exp_lines", []):
    who = e.get("paid_to_name") or ("ผู้บริหาร" if e.get("paid_to_type") == "manager" else "สมาชิก")
    detail(f'{e.get("description") or "จ่ายพิเศษ"} → {who}', e["amount"])

# ค่าบริหารจัดการทั่วไป
line("ค่าใช้จ่ายบริหารจัดการ", d["general_exp_total"], color=RED, sign="-")
for e in d.get("general_exp_lines", []):
    detail(e.get("description"), e["amount"])

# ── กำไรสุทธิ ──
row += 1
cell(f"B{row}", "กำไรสุทธิสุดท้าย (Net Profit)", font=Font(name=FONT, size=12, bold=True, color="FFFFFF"),
     fill=(GREEN if d["net"] >= 0 else RED), align=L)
cell(f"C{row}", None, fill=(GREEN if d["net"] >= 0 else RED))
cell(f"D{row}", d["net"], font=Font(name=FONT, size=12, bold=True, color="FFFFFF"),
     fill=(GREEN if d["net"] >= 0 else RED), align=R, fmt=MONEY)
ws.row_dimensions[row].height = 26
margin = (d["net"] / d["revenue"] * 100) if d["revenue"] else 0
row += 1
ws.merge_cells(f"B{row}:D{row}")
cell(f"B{row}", f"อัตรากำไรสุทธิ {margin:.1f}%  ·  กำไรขั้นต้น {d['gross']:,.2f} บาท", font=Font(name=FONT, size=9.5, italic=True, color=GREY), align=R)

# print setup
ws.print_area = f"A1:E{row+1}"
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.4

wb.save(out)
print(out)
