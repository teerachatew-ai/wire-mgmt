# -*- coding: utf-8 -*-
# Export ตารางตรวจสอบสต็อค (Check & Balance) เป็นไฟล์ Excel
# Usage: python stock_export.py <data.json> <out.xlsx>
import sys, json, warnings
warnings.simplefilter("ignore")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

dataf, out = sys.argv[1], sys.argv[2]
d = json.load(open(dataf, encoding="utf-8-sig"))
rows = d.get("products", [])
month = d.get("month")

TH_MONTH = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
            "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]

FONT = "Tahoma"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "ตรวจสอบสต็อค"
ws.sheet_view.showGridLines = False

# คอลัมน์ตามโหมด (มีเดือน = แสดงยกมา/ยกไป)
if month:
    y, mm = month.split("-")
    subtitle = f"เดือน {TH_MONTH[int(mm)]} {int(y)+543} (แสดงเฉพาะยอดเคลื่อนไหวในเดือน)"
    cols = [("สินค้า", "name", "text"), ("ยกมาจากเดือนก่อนหน้า", "carry_ready", "n"), ("รับเข้า", "received", "n"),
            ("เบิกออก", "total_issued", "n"), ("งานรอแจกจ่าย", "wait_distribute", "n"), ("งานดี", "ret_good", "n"),
            ("NG จากการตัด", "ret_ngcut", "n"), ("NG จากโรงงาน", "ret_ngfac", "n"),
            ("ยอดส่งโรงงาน", "shipped", "n"), ("ยอดยกไปเดือนถัดไป", "closing_ready", "n")]
else:
    subtitle = "ยอดสะสมทั้งหมด"
    cols = [("สินค้า", "name", "text"), ("รับเข้า (สะสม)", "received", "n"), ("เบิกออก", "total_issued", "n"),
            ("คืนดี", "ret_good", "n"), ("คืนเสีย", "ret_defect", "n"), ("เศษ", "ret_waste", "n"),
            ("ส่งออก", "shipped", "n"), ("พร้อมส่ง (คงเหลือ)", "available", "n")]

ncol = len(cols)
last_col = get_column_letter(ncol)

# หัวเรื่อง
ws.merge_cells(f"A1:{last_col}1")
c = ws["A1"]; c.value = "ภาพรวมสต๊อค & Check Balance"
c.font = Font(name=FONT, size=15, bold=True, color="1F2937"); c.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 26
ws.merge_cells(f"A2:{last_col}2")
c = ws["A2"]; c.value = subtitle
c.font = Font(name=FONT, size=10, color="6B7280")
ws.row_dimensions[2].height = 18

# หัวตาราง (แถว 4)
HR = 4
head_fill = PatternFill("solid", fgColor="374151")
for i, (label, _, _) in enumerate(cols, start=1):
    cell = ws.cell(row=HR, column=i, value=label)
    cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    cell.fill = head_fill
    cell.alignment = Alignment(horizontal=("left" if i == 1 else "right"), vertical="center")
    cell.border = border
ws.row_dimensions[HR].height = 22

# ข้อมูล
NUMFMT = "#,##0"
r = HR + 1
totals = {key: 0 for _, key, t in cols if t == "n"}
for p in rows:
    for i, (label, key, t) in enumerate(cols, start=1):
        cell = ws.cell(row=r, column=i)
        cell.border = border
        if t == "text":
            unit = p.get("unit") or ""
            cell.value = f'{p.get("code","")}  {p.get("name","")}' + (f" ({unit})" if unit else "")
            cell.font = Font(name=FONT, size=10.5, color="111827")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            v = p.get(key)
            v = 0 if v is None else v
            cell.value = v
            cell.number_format = NUMFMT
            cell.font = Font(name=FONT, size=10.5, color="111827")
            cell.alignment = Alignment(horizontal="right", vertical="center")
            totals[key] += v
    ws.row_dimensions[r].height = 19
    r += 1

# แถวรวม
tot_fill = PatternFill("solid", fgColor="ECFDF5")
for i, (label, key, t) in enumerate(cols, start=1):
    cell = ws.cell(row=r, column=i)
    cell.fill = tot_fill; cell.border = border
    if i == 1:
        cell.value = "รวม"
        cell.font = Font(name=FONT, size=11, bold=True, color="065F46")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    else:
        cell.value = totals.get(key, 0)
        cell.number_format = NUMFMT
        cell.font = Font(name=FONT, size=11, bold=True, color="065F46")
        cell.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[r].height = 22

# ความกว้างคอลัมน์
ws.column_dimensions["A"].width = 34
for i in range(2, ncol + 1):
    ws.column_dimensions[get_column_letter(i)].width = 13

# ตรึงหัวตาราง + print setup
ws.freeze_panes = f"A{HR+1}"
ws.print_area = f"A1:{last_col}{r}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(out)
print(out)
