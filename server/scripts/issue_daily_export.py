# -*- coding: utf-8 -*-
# รายงานใบเบิกงานรายวัน — 1 วัน = 1 ชีต (แปลงเป็น PDF แล้วแต่ละชีตกลายเป็นหน้าเรียงต่อกันตามลำดับวันที่)
# แถว = สมาชิกแต่ละคน, คอลัมน์ = ชนิดสินค้า (matrix) + subtotal ท้ายตาราง + ยอดรับเข้าจากโรงงานวันนั้น
# Usage: python issue_daily_export.py <data.json> <out.xlsx>
import sys, json, re, warnings
warnings.simplefilter("ignore")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

dataf, out = sys.argv[1], sys.argv[2]
d = json.load(open(dataf, encoding="utf-8-sig"))

TH = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
      "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]

def date_th(iso):
    if not iso:
        return "-"
    s = str(iso)[:10]
    parts = s.split("-")
    if len(parts) != 3:
        return s
    y, m, dd = parts
    return f"{dd} {TH[int(m)]} {int(y) + 543}"

def hexcolor(c):
    if not c:
        return None
    c = str(c).lstrip("#").strip()
    if len(c) == 3:
        c = "".join(ch * 2 for ch in c)
    if len(c) == 6 and all(ch in "0123456789abcdefABCDEF" for ch in c):
        return c.upper()
    return None

def contrast_text(hexc):
    if not hexc:
        return "111827"
    r, g, b = int(hexc[0:2], 16), int(hexc[2:4], 16), int(hexc[4:6], 16)
    lum = 0.299 * r + 0.587 * g + 0.114 * b
    return "111827" if lum > 160 else "FFFFFF"

def short_label(name):
    mm = re.search(r"\(([^)]+)\)", name or "")
    return mm.group(1) if mm else (name or "-")

def code_num(name):
    prefix = (name or "").split(" (")[0].strip()
    mm = re.search(r'-(\d+)', prefix)
    if mm:
        return mm.group(1)
    mm = re.search(r'(\d+)', prefix)
    return mm.group(1) if mm else ""

FONT = "Tahoma"
NAVY = "1E3A5F"; GREEN = "0B7A3B"; RED = "B42318"; GREY = "6B7280"; AMBER = "B45309"; SKY = "0369A1"
NUM = '#,##0'
NUM_Z = '#,##0;-#,##0;"-"'
thin = Side(style="thin", color="D8DEE9")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)

def safe_sheet_name(name, used):
    name = re.sub(r'[\\/*?:\[\]]', ' ', name)[:28].strip() or "sheet"
    base, i = name, 2
    while name in used:
        name = f"{base}-{i}"; i += 1
    used.add(name)
    return name

def cell(ws, coord, val, *, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws[coord]; c.value = val
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = border
    return c

R = Alignment(horizontal="right", vertical="center")
L = Alignment(horizontal="left", vertical="center")
C = Alignment(horizontal="center", vertical="center")
CW = Alignment(horizontal="center", vertical="center", wrap_text=True)
LW = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── รวบรวมรายชื่อสินค้าทั้งหมดที่ปรากฏในช่วงที่ export จัดกลุ่มตามสี (สีเดียวกันอยู่ติดกัน) ──
distinct_products = {name: info.get("color") for name, info in d.get("products", {}).items()}

def color_priority(hexc):
    if not hexc:
        return 9
    r, g, b = int(hexc[0:2], 16), int(hexc[2:4], 16), int(hexc[4:6], 16)
    if r > 200 and g > 200 and b > 200:
        return 0
    if g > r and g > b:
        return 2
    if r >= g and r >= b:
        return 1
    return 3

def color_sort_key(name):
    hexc = hexcolor(distinct_products[name])
    return (color_priority(hexc), hexc or "ZZZZZZ", code_num(name), name)

product_order = sorted(distinct_products.keys(), key=color_sort_key)

def base_label(name):
    lbl = short_label(name)
    code = code_num(name)
    return f"{lbl} {code}".strip() if code else lbl

label_freq = {}
for name in product_order:
    lbl = base_label(name)
    label_freq[lbl] = label_freq.get(lbl, 0) + 1
product_label = {}
for name in product_order:
    lbl = base_label(name)
    if label_freq[lbl] > 1:
        prefix = name.split(" (")[0].strip()
        lbl = f"{lbl} ({prefix})"
    product_label[name] = lbl

n_prod = len(product_order)
FIXED_COLS = 3  # รหัส, ชื่อ-สกุล, ชื่อเล่น
LAST_COL = FIXED_COLS + n_prod + 1  # + สินค้าแต่ละชนิด + รวม
LAST_LETTER = get_column_letter(LAST_COL)
TOTAL_LETTER = LAST_LETTER
PROD_START_LETTER = get_column_letter(FIXED_COLS + 1)

used_names = set()

for day in d["days"]:
    ws = wb.create_sheet(safe_sheet_name(day["date"], used_names))
    ws.sheet_view.showGridLines = False

    widths = [9, 22, 13] + [11] * n_prod + [11]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.merge_cells(f"A1:{LAST_LETTER}1")
    cell(ws, "A1", d.get("org_name", ""), font=Font(name=FONT, size=13, bold=True, color=NAVY), align=CW)
    ws.merge_cells(f"A2:{LAST_LETTER}2")
    cell(ws, "A2", f'รายงานใบเบิกงานรายวัน — วันที่ {date_th(day["date"])}', font=Font(name=FONT, size=11, color=GREY), align=CW)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24

    member_count = day.get("member_count", len(day["members"]))
    ws.merge_cells(f"A3:{LAST_LETTER}3")
    cell(ws, "A3", f'จำนวนสมาชิกที่เบิกงานวันนี้ {member_count} คน', font=Font(name=FONT, size=10.5, bold=True, color=SKY), align=CW)
    ws.row_dimensions[3].height = 20

    row = 4
    headers = ["รหัส", "ชื่อ-สกุล", "ชื่อเล่น"] + [product_label[n] for n in product_order] + ["รวม"]
    for ci, h in enumerate(headers, start=1):
        col = get_column_letter(ci)
        is_prod_col = FIXED_COLS + 1 <= ci <= FIXED_COLS + n_prod
        if is_prod_col:
            pname = product_order[ci - FIXED_COLS - 1]
            hexc = hexcolor(distinct_products[pname]) or "9CA3AF"
            fill, txt = hexc, contrast_text(hexc)
        else:
            fill, txt = NAVY, "FFFFFF"
        cell(ws, f"{col}{row}", h, font=Font(name=FONT, size=9.5, bold=True, color=txt), fill=fill, align=C, border=box)
    ws.row_dimensions[row].height = 24
    row += 1

    col_totals = {n: 0 for n in product_order}
    grand_total = 0
    for m in day["members"]:
        vals = [m["member_code"], m["member_name"], m.get("member_nickname") or "-"]
        vals += [m["qty"].get(n, 0) for n in product_order]
        vals += [m["total"]]
        for ci, v in enumerate(vals, start=1):
            col = get_column_letter(ci)
            is_prod_col = FIXED_COLS + 1 <= ci <= FIXED_COLS + n_prod
            is_total_col = ci == LAST_COL
            fmt = NUM_Z if is_prod_col else (NUM if is_total_col else None)
            cell(ws, f"{col}{row}", v, font=Font(name=FONT, size=9.5, bold=is_total_col, color="111827"),
                 align=(R if (is_prod_col or is_total_col) else L), border=box, fmt=fmt)
        for n in product_order:
            col_totals[n] += m["qty"].get(n, 0)
        grand_total += m["total"]
        ws.row_dimensions[row].height = 17
        row += 1

    # ── บรรทัดรวม (subtotal) ต่อคอลัมน์ ──
    ws.merge_cells(f"A{row}:C{row}")
    cell(ws, f"A{row}", "รวมเบิกวันนี้", font=Font(name=FONT, size=10, bold=True, color="FFFFFF"), fill=GREEN, align=R, border=box)
    for ci, n in enumerate(product_order, start=FIXED_COLS + 1):
        col = get_column_letter(ci)
        cell(ws, f"{col}{row}", col_totals.get(n, 0), font=Font(name=FONT, size=10, bold=True, color="FFFFFF"), fill=GREEN, align=R, fmt=NUM_Z, border=box)
    cell(ws, f"{LAST_LETTER}{row}", grand_total, font=Font(name=FONT, size=10, bold=True, color="FFFFFF"), fill=GREEN, align=R, fmt=NUM, border=box)
    ws.row_dimensions[row].height = 20
    row += 2

    # ── ยอดรับเข้าจากโรงงานวันนี้ (ถ้ามี) ──
    receives = day.get("receives") or {}
    if receives:
        ws.merge_cells(f"A{row}:C{row}")
        cell(ws, f"A{row}", "📦 รับเข้าจากโรงงานวันนี้", font=Font(name=FONT, size=10, bold=True, color="FFFFFF"), fill=SKY, align=R, border=box)
        recv_total = 0
        for ci, n in enumerate(product_order, start=FIXED_COLS + 1):
            col = get_column_letter(ci)
            qty = receives.get(n, 0)
            recv_total += qty
            cell(ws, f"{col}{row}", (qty if qty else None), font=Font(name=FONT, size=10, bold=True, color="FFFFFF"), fill=SKY, align=R, fmt=NUM_Z, border=box)
        cell(ws, f"{LAST_LETTER}{row}", recv_total, font=Font(name=FONT, size=10, bold=True, color="FFFFFF"), fill=SKY, align=R, fmt=NUM, border=box)
        ws.row_dimensions[row].height = 20
        row += 1
    else:
        ws.merge_cells(f"A{row}:{LAST_LETTER}{row}")
        cell(ws, f"A{row}", "🚫 วันนี้ไม่มียอดรับเข้าจากโรงงาน", font=Font(name=FONT, size=10, italic=True, bold=True, color=AMBER), align=LW)
        ws.row_dimensions[row].height = 22
        row += 1

    ws.print_area = f"A1:{LAST_LETTER}{row}"
    ws.page_setup.orientation = "portrait" if n_prod <= 4 else "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.4

wb.save(out)
print(out)
