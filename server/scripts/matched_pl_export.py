# -*- coding: utf-8 -*-
# งบการเงิน (แบบจับคู่ต้นทุน-รายรับ) รายเดือน — แยกต่างหากจาก pl_export.py เดิมโดยสิ้นเชิง
# Usage: python matched_pl_export.py <data.json> <out.xlsx>
import sys, json, warnings
warnings.simplefilter("ignore")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

dataf, out = sys.argv[1], sys.argv[2]
d = json.load(open(dataf, encoding="utf-8-sig"))

TH = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
      "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
y, m = d["month"].split("-")
month_th = f"{TH[int(m)]} {int(y) + 543}"

FONT = "Tahoma"
NAVY = "1E3A5F"; GREEN = "0B7A3B"; RED = "B42318"; GREY = "6B7280"; AMBER = "B45309"
MONEY = '#,##0.00;[Red](#,##0.00)'
thin = Side(style="thin", color="D8DEE9")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "งบการเงิน"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 46
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 3

def cell(coord, val, *, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws[coord]; c.value = val
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = border
    return c

R = Alignment(horizontal="right", vertical="center")
L = Alignment(horizontal="left", vertical="center", wrap_text=True)
C = Alignment(horizontal="center", vertical="center")

# ── หัวรายงาน ──
ws.merge_cells("B2:D2")
cell("B2", d.get("org_name", ""), font=Font(name=FONT, size=13, bold=True, color=NAVY), align=C)
ws.merge_cells("B3:D3")
cell("B3", f"งบการเงิน (แบบจับคู่ต้นทุน-รายรับ)  ประจำเดือน {month_th}", font=Font(name=FONT, size=11, color=GREY), align=C)
ws.merge_cells("B4:D4")
cell("B4", "ต้นทุนค่าแรงคิดจากปริมาณที่ส่งออก/วางบิลจริงเดือนนี้เท่านั้น — ตรงตามหลักการจับคู่ต้นทุน-รายรับ (matching principle)",
     font=Font(name=FONT, size=9, italic=True, color=GREY), align=C)
ws.row_dimensions[2].height = 22

row = 6
def section(title):
    global row
    ws.merge_cells(f"B{row}:D{row}")
    cell(f"B{row}", title, font=Font(name=FONT, size=10.5, bold=True, color="FFFFFF"), fill=NAVY, align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[row].height = 20
    row += 1

def line(label, amount, *, bold=False, color=None, sign=""):
    global row
    cell(f"B{row}", label, font=Font(name=FONT, size=10.5, bold=bold, color=color or "111827"), align=Alignment(horizontal="left", vertical="center"), border=box)
    cell(f"C{row}", None, border=box)
    disp = amount if sign != "-" else -abs(amount)
    cell(f"D{row}", disp, font=Font(name=FONT, size=10.5, bold=bold, color=color or "111827"),
         align=R, fmt=MONEY, border=box)
    ws.row_dimensions[row].height = 18
    row += 1

def detail(label, amount):
    global row
    cell(f"B{row}", "      • " + (label or "(ไม่ระบุ)"), font=Font(name=FONT, size=9.5, color=GREY), align=Alignment(horizontal="left", vertical="center"), border=box)
    cell(f"C{row}", amount, font=Font(name=FONT, size=9.5, color=GREY), align=R, fmt=MONEY, border=box)
    cell(f"D{row}", None, border=box)
    ws.row_dimensions[row].height = 16
    row += 1

def note(text, *, color=GREY):
    global row
    ws.merge_cells(f"B{row}:D{row}")
    cell(f"B{row}", text, font=Font(name=FONT, size=9, italic=True, color=color), align=L)
    ws.row_dimensions[row].height = 26
    row += 1

# ── รายรับ ──
section("รายรับ")
line("รายรับจากโรงงาน (Amphenol) — ตามยอดส่งออก/วางบิลจริงเดือนนี้", d["revenue"], bold=True, color=GREEN)

# ── ต้นทุนขาย (COGS) ──
section("หัก ต้นทุนขาย (ค่าแรงของงานที่ส่งออกจริงเดือนนี้)")
line("ค่าแรงสมาชิก (COGS — เฉพาะของที่ส่งออกเดือนนี้)", d["cogs"], color=RED, sign="-")
row += 1
cell(f"B{row}", "กำไรขั้นต้น (Gross Profit)", font=Font(name=FONT, size=11, bold=True, color=NAVY), align=Alignment(horizontal="left", vertical="center"))
cell(f"D{row}", d["gross"], font=Font(name=FONT, size=11, bold=True, color=NAVY), align=R, fmt=MONEY)
ws.row_dimensions[row].height = 20
row += 2

# ── ค่าใช้จ่ายดำเนินงาน ──
section("หัก ค่าใช้จ่ายดำเนินงาน")
line(f"ภาษี ณ ที่จ่าย {d.get('tax_pct', 3)}%", d["tax"], color=RED, sign="-")

for mg in d.get("manager_lines", []):
    if mg["computed"]:
        detail(f'{mg["name"]}{(" · " + mg["role"]) if mg.get("role") else ""}', mg["computed"])
for e in d.get("comp_exp_lines", []):
    who = e.get("paid_to_name") or ("ผู้บริหาร" if e.get("paid_to_type") == "manager" else "สมาชิก")
    detail(f'{e.get("description") or "จ่ายพิเศษ"} → {who}', e["amount"])
line("รวมค่าตอบแทนผู้บริหาร", d["manager_comp"], color=RED, sign="-", bold=True)

for e in d.get("general_exp_lines", []):
    detail(e.get("description"), e["amount"])
line("รวมค่าใช้จ่ายบริหารจัดการ", d["general_exp_total"], color=RED, sign="-", bold=True)

# ── กำไรสุทธิ (matched) ──
row += 1
cell(f"B{row}", "กำไรสุทธิ (แบบจับคู่ต้นทุน-รายรับ)", font=Font(name=FONT, size=12, bold=True, color="FFFFFF"),
     fill=(GREEN if d["net_matched"] >= 0 else RED), align=Alignment(horizontal="left", vertical="center"))
cell(f"C{row}", None, fill=(GREEN if d["net_matched"] >= 0 else RED))
cell(f"D{row}", d["net_matched"], font=Font(name=FONT, size=12, bold=True, color="FFFFFF"),
     fill=(GREEN if d["net_matched"] >= 0 else RED), align=R, fmt=MONEY)
ws.row_dimensions[row].height = 26
margin = (d["net_matched"] / d["revenue"] * 100) if d["revenue"] else 0
row += 1
ws.merge_cells(f"B{row}:D{row}")
cell(f"B{row}", f"อัตรากำไรสุทธิ {margin:.1f}%", font=Font(name=FONT, size=9.5, italic=True, color=GREY), align=R)
row += 2

# ── เทียบกับวิธีคิดแบบเดิม (ตามรอบจ่ายค่าแรง) ──
section("เทียบกับวิธีคิดแบบเดิม (ตามรอบจ่ายค่าแรง)")
line("ค่าแรงสมาชิกตามรอบจ่าย (จ่ายจริงเดือนนี้)", d["cash_basis_wage"], color=GREY)
line("กำไรสุทธิแบบเดิม (ตามรอบจ่ายค่าแรง)", d["net_cash_basis"], color=(GREEN if d["net_cash_basis"] >= 0 else RED), bold=True)
line("ส่วนต่าง (แบบเดิม − แบบจับคู่)", d["variance"], color=AMBER, bold=True)
note("ส่วนต่างเกิดจากค่าแรงตามรอบจ่ายรวมงานของเดือนอื่นปนอยู่ (ตัดเสร็จเดือนนี้แต่ยังไม่ส่งออก หรือส่งออกเดือนนี้แต่ตัดเสร็จเดือนก่อน) "
     "ทำให้ไม่ตรงกับรายรับที่รับรู้เดือนนี้ — ตัวเลข \"แบบจับคู่\" ด้านบนคือกำไรที่แท้จริงของเดือนนี้")
row += 1

# ── สินทรัพย์/หนี้สินที่เกี่ยวข้อง ณ สิ้นเดือน ──
section("สินทรัพย์-หนี้สินที่เกี่ยวข้องกับงานตัดสายไฟ ณ สิ้นเดือน")
line("สินค้าคงเหลือ (งานตัดเสร็จ รอส่งออก) ยกมา", d["inventory_open"], color=GREY)
line("สินค้าคงเหลือ (งานตัดเสร็จ รอส่งออก) ยกไป", d["inventory_close"], color=NAVY, bold=True)
line("ค่าแรงค้างจ่าย (จ่ายจริงวันที่ 25 ของเดือนถัดไป)", d["accrued_wages_payable"], color=RED, bold=True)
note("รายการนี้เฉพาะส่วนที่เกี่ยวข้องกับงานตัดสายไฟที่ระบบติดตามได้ ไม่ใช่งบดุลฉบับเต็ม (ยังไม่รวมเงินสด/บัญชีธนาคาร/ทุน)", color=AMBER)

# print setup
ws.print_area = f"A1:E{row+1}"
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.4

wb.save(out)
print(out)
