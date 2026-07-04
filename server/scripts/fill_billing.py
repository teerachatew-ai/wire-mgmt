# -*- coding: utf-8 -*-
# Fills the real billing template (server/templates/billing-template.xlsx) with data.
# Produces a 2-page billing note:
#   Page 1 = "สรุป Summary"  → รวมยอดต่อประเภทสินค้า (group เหมือนใบแจ้งหนี้)
#   Page 2 = "รายวัน Detail" → เรียงตามวันที่ (รายการละเอียดเหมือนเดิม)
# Usage: python fill_billing.py <template.xlsx> <data.json> <out.xlsx> [pdf]
import sys, json, datetime, warnings
from copy import copy
warnings.simplefilter("ignore")
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.properties import PageSetupProperties

tpl, dataf, out = sys.argv[1], sys.argv[2], sys.argv[3]
mode = sys.argv[4] if len(sys.argv) > 4 else ""   # "pdf" = เหลือเฉพาะชีตฟอร์ม
d = json.load(open(dataf, encoding="utf-8-sig"))

wb = load_workbook(tpl)
FORM_SHEET = "สำหรับกรอกข้อมูล Form "
SUMMARY_TITLE = "สรุป Summary"
DETAIL_TITLE  = "รายวัน Detail"

month = d.get("month", "")  # 'YYYY-MM'
sup = d.get("supplier", {})
wht_rate = d.get("wht_rate", 0.03)
lines = d.get("lines", [])

TH = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
      "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
EN = ["January", "February", "March", "April", "May", "June",
      "July", "August", "September", "October", "November", "December"]

BASE = 17          # first data row
CAP = 26           # template provides rows 17..42
MINROWS = 40       # จำนวนแถวมาตรฐานในฟอร์ม


# ── สรุปรวมต่อประเภทสินค้า (group by project+part+description+price) ─────────
def summarize(rows):
    groups = {}
    order = []
    for l in rows:
        key = (l.get("project") or "", l.get("part_number") or "", l.get("description") or "", l.get("price") or 0)
        if key not in groups:
            groups[key] = {"project": l.get("project"), "part_number": l.get("part_number"),
                           "description": l.get("description"), "price": l.get("price"),
                           "quantity": 0, "deliveryDate": None}
            order.append(key)
        groups[key]["quantity"] += (l.get("quantity") or 0)
    out_rows = [groups[k] for k in order]
    out_rows.sort(key=lambda r: (str(r.get("project") or ""), str(r.get("part_number") or "")))
    return out_rows


# ── กรอกข้อมูลลงชีตฟอร์มหนึ่งชีต (header + supplier + rows + footer + page) ──
def fill_form(ws, rows):
    # หัวคอลัมน์ C: PO -> ชื่อโครงการ
    ws["C15"] = "ชื่อโครงการ"
    ws["C16"] = "Project"

    if month:
        y, m = month.split("-")
        ws["F5"] = f"{TH[int(m)-1]} {EN[int(m)-1]}"
        ws["I5"] = int(y)

    if sup.get("name"):    ws["D8"]  = sup["name"]
    if sup.get("code"):
        ws["H8"] = sup["code"]
        _a = ws["H8"].alignment
        ws["H8"].alignment = Alignment(horizontal="center", vertical=_a.vertical or "center")
    if sup.get("address"): ws["D10"] = sup["address"]
    if sup.get("contact"): ws["D12"] = sup["contact"]
    if sup.get("tel"):     ws["H12"] = sup["tel"]

    # วันที่บนใบวางบิล = วันที่ export (วันนี้) — เขียนก่อน insert เพื่อให้เลื่อนตาม
    ws["C50"] = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    # ป้าย "ผู้วางบิล"/"ผู้รับวางบิล" ล้นชนช่องเซ็น -> หดให้พอดี
    for coord in ("B48", "G48"):
        c = ws[coord]
        a = c.alignment
        c.alignment = Alignment(horizontal=a.horizontal, vertical=a.vertical, shrink_to_fit=True)

    n = len(rows)
    target = max(MINROWS, n)

    def copy_row_style(src, dst):
        for col in range(1, 13):  # A..L
            s = ws.cell(row=src, column=col)
            t = ws.cell(row=dst, column=col)
            t._style = copy(s._style)

    DATA_H = ws.row_dimensions[17].height or 19.5

    if target > CAP:
        extra = target - CAP
        foot_merges = [str(mm) for mm in ws.merged_cells.ranges if mm.min_row >= 43]
        foot_heights = {r: ws.row_dimensions[r].height for r in range(43, 70) if ws.row_dimensions[r].height is not None}
        for rng in foot_merges:
            ws.unmerge_cells(rng)
        ws.insert_rows(43, extra)
        for rng in foot_merges:
            c1, r1, c2, r2 = range_boundaries(rng)
            ws.merge_cells(start_row=r1 + extra, end_row=r2 + extra, start_column=c1, end_column=c2)
        for r, h in foot_heights.items():
            ws.row_dimensions[r + extra].height = h
        for i in range(extra):
            r = 42 + 1 + i
            copy_row_style(17, r)
            ws.cell(row=r, column=2).value = "=ROW()-16"
            ws.cell(row=r, column=9).value  = f'=IF(F{r}*H{r}=0,"",F{r}*H{r})'
            ws.cell(row=r, column=10).value = f'=IF(I{r}="","",I{r}*K{r})'
            ws.cell(row=r, column=11).value = 0.03

    tr = BASE + target
    last = tr - 1

    for idx in range(BASE, last + 1):
        li = idx - BASE
        copy_row_style(BASE, idx)
        ws.row_dimensions[idx].height = DATA_H
        ws.cell(row=idx, column=2).value  = "=ROW()-16"
        ws.cell(row=idx, column=9).value  = f'=IF(F{idx}*H{idx}=0,"",F{idx}*H{idx})'
        ws.cell(row=idx, column=10).value = f'=IF(I{idx}="","",I{idx}*K{idx})'
        if li < n:
            l = rows[li]
            ws.cell(row=idx, column=3).value = l.get("project") or None
            ws.cell(row=idx, column=4).value = l.get("part_number") or None
            ws.cell(row=idx, column=5).value = l.get("description") or None
            ws.cell(row=idx, column=6).value = l.get("quantity") or None
            dt = l.get("deliveryDate")
            if dt:
                try:
                    ws.cell(row=idx, column=7).value = datetime.datetime.strptime(dt[:10], "%Y-%m-%d")
                except Exception:
                    ws.cell(row=idx, column=7).value = dt
            else:
                ws.cell(row=idx, column=7).value = None
            ws.cell(row=idx, column=8).value = l.get("price") or None
            ws.cell(row=idx, column=11).value = wht_rate
        else:
            for col in (3, 4, 5, 6, 7, 8, 11):
                ws.cell(row=idx, column=col).value = None

    ws.cell(row=tr, column=6).value  = f'=IF(SUMIF(F{BASE}:F{last},"<>")=0,"",SUMIF(F{BASE}:F{last},"<>"))'
    ws.cell(row=tr, column=9).value  = f'=IF(SUM(I{BASE}:I{last})="","",SUM(I{BASE}:I{last}))'
    ws.cell(row=tr, column=10).value = f'=IF(SUM(J{BASE}:J{last})="","",SUM(J{BASE}:J{last}))'

    nr = tr + 2
    ws.cell(row=nr, column=10).value = f'=IF(J{tr}="","",I{tr}-J{tr})'

    extra2 = max(0, target - CAP)
    footer_last = 63 + extra2
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None
    ws.print_area = f"A1:L{footer_last}"
    ws.print_title_rows = None


# ── สร้าง 2 ชีต: สรุป (หน้า 1) + รายวัน (หน้า 2) ────────────────────────────
ws_detail = wb[FORM_SHEET]
ws_summary = wb.copy_worksheet(ws_detail)   # คัดลอกจากเทมเพลตต้นฉบับ (ยังไม่แก้)
ws_summary.title = SUMMARY_TITLE
ws_detail.title = DETAIL_TITLE

# กรอกข้อมูล
fill_form(ws_summary, summarize(lines))
fill_form(ws_detail, lines)

# จัดลำดับ: สรุปไว้หน้าแรก, รายวันหน้าถัดไป
wb.move_sheet(ws_summary, -(wb.index(ws_summary)))              # ย้ายสรุปไปตำแหน่งแรกสุด
idx_sum = wb.index(ws_summary)
wb.move_sheet(ws_detail, (idx_sum + 1) - wb.index(ws_detail))  # รายวันต่อจากสรุป

# โหมด PDF: เหลือเฉพาะ 2 ชีตฟอร์ม (ตัดหน้าตัวอย่าง/เดือน ออก)
if mode == "pdf":
    for sn in list(wb.sheetnames):
        if sn not in (SUMMARY_TITLE, DETAIL_TITLE):
            del wb[sn]

wb.active = wb.index(ws_summary)
wb.save(out)
print(out)
