# -*- coding: utf-8 -*-
# รายงานเบิกงาน/ส่งงานรายบุคคล ประจำรอบจ่ายค่าแรง — 1 คน = 1 ชีต, รวมทุกคนในไฟล์เดียว
# แปลงเป็น PDF แล้วแต่ละชีตจะกลายเป็นหน้าเรียงต่อกันตามลำดับโดยอัตโนมัติ
# Usage: python payroll_detail_export.py <data.json> <out.xlsx>
import sys, json, re, warnings
warnings.simplefilter("ignore")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

dataf, out = sys.argv[1], sys.argv[2]
d = json.load(open(dataf, encoding="utf-8-sig"))

# ขนาดกระดาษเลือกได้ A4/A5 — A5 บังคับแนวนอน + เต็มหน้า + กึ่งกลาง (พิมพ์ใบเล็กแล้วยังอ่านง่าย)
PAPER_SIZE_CODE = {"A4": "9", "A5": "11"}
paper_size = d.get("paper_size") if d.get("paper_size") in PAPER_SIZE_CODE else "A4"
duplicate_for_pdf = bool(d.get("duplicate_for_pdf"))
# print เต็มรูปแบบ (ต้นฉบับ+คู่ฉบับ) หรือแค่ตรวจทาน (ต้นฉบับอย่างเดียว) — มีผลเฉพาะตอน duplicate_for_pdf เท่านั้น
include_copy = bool(d.get("include_copy", True))

TH = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
      "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]

def month_th(ym):
    y, m = ym.split("-")
    return f"{TH[int(m)]} {int(y) + 543}"

def date_th(iso):
    if not iso:
        return "-"
    s = str(iso)[:10]
    parts = s.split("-")
    if len(parts) != 3:
        return s
    y, m, dd = parts
    return f"{dd}/{m}/{int(y) + 543}"

def hexcolor(c):
    if not c:
        return None
    c = str(c).lstrip("#").strip()
    if len(c) == 3:
        c = "".join(ch * 2 for ch in c)
    if len(c) == 6 and all(ch in "0123456789abcdefABCDEF" for ch in c):
        return c.upper()
    return None

def contrast_text(hexc):
    if not hexc:
        return "111827"
    r, g, b = int(hexc[0:2], 16), int(hexc[2:4], 16), int(hexc[4:6], 16)
    lum = 0.299 * r + 0.587 * g + 0.114 * b
    return "111827" if lum > 160 else "FFFFFF"

def short_label(name):
    mm = re.search(r"\(([^)]+)\)", name or "")
    return mm.group(1) if mm else (name or "-")

FONT = "Cordia New"
FONT_SCALE = 1.2  # ขยายตัวหนังสือทุกจุดในรายงานขึ้น — ค่าสูงสุดที่ทดสอบแล้วไม่ตกขอบ/ล้นหน้ากระดาษ
def FS(size):
    return round(size * FONT_SCALE, 2)
def RH(height):  # ขยายความสูงแถวตามสัดส่วนฟอนต์ ป้องกันตัวหนังสือถูกตัด
    return round(height * FONT_SCALE, 1)
NAVY = "1E3A5F"; GREEN = "0B7A3B"; RED = "B42318"; GREY = "6B7280"; AMBER = "B45309"
NUM = '#,##0'
NUM_Z = '#,##0;-#,##0;"-"'
MONEY = '#,##0.00;[Red](#,##0.00)'
MONEY_Z = '#,##0.00;[Red](#,##0.00);"-"'
thin = Side(style="thin", color="D8DEE9")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)

def safe_sheet_name(name, used):
    name = re.sub(r'[\\/*?:\[\]]', ' ', name)[:28].strip() or "sheet"
    base, i = name, 2
    while name in used:
        name = f"{base}-{i}"; i += 1
    used.add(name)
    return name

def cell(ws, coord, val, *, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws[coord]; c.value = val
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = border
    return c

R = Alignment(horizontal="right", vertical="center")
L = Alignment(horizontal="left", vertical="center")
C = Alignment(horizontal="center", vertical="center")
CW = Alignment(horizontal="center", vertical="center", wrap_text=True)
LW = Alignment(horizontal="left", vertical="center", wrap_text=True)
RW = Alignment(horizontal="right", vertical="center", wrap_text=True)

used_names = set()

def code_num(name):
    prefix = (name or "").split(" (")[0].strip()
    mm = re.search(r'-(\d+)', prefix)
    if mm:
        return mm.group(1)
    mm = re.search(r'(\d+)', prefix)
    return mm.group(1) if mm else ""

# ── รวบรวมรายชื่อสินค้าทั้งหมด จัดกลุ่มตามสี (สีเดียวกันอยู่ติดกัน) แล้วเรียงตามรหัสในกลุ่มนั้น ──
distinct_products = {}
for m in d["members"]:
    for pw in m.get("product_wages", []):
        distinct_products.setdefault(pw["name"], pw.get("color"))

def color_priority(hexc):
    # ลำดับที่ต้องการ: ขาวก่อน -> ชมพู/แดง -> เขียวไว้ขวาสุด -> สีอื่นๆ
    if not hexc:
        return 9
    r, g, b = int(hexc[0:2], 16), int(hexc[2:4], 16), int(hexc[4:6], 16)
    if r > 200 and g > 200 and b > 200:
        return 0
    if g > r and g > b:
        return 2
    if r >= g and r >= b:
        return 1
    return 3

def color_sort_key(name):
    hexc = hexcolor(distinct_products[name])
    return (color_priority(hexc), hexc or "ZZZZZZ", code_num(name), name)

product_order = sorted(distinct_products.keys(), key=color_sort_key)

# ตัดคำว่า "ป้าย"/"เส้น" ออกจากหัวคอลัมน์ — คำเหล่านี้มีทุกชื่อสินค้าอยู่แล้วไม่ช่วยแยกแยะ
# แต่ทำให้ข้อความยาวจนล้นช่องตาราง (ไม่ลดขนาดฟอนต์ ตัดคำแทน)
def strip_noise_words(s):
    return s.replace("ป้าย", "").replace("เส้น", "").strip()

def base_label(name):
    lbl = strip_noise_words(short_label(name))
    code = code_num(name)
    return f"{lbl} {code}".strip() if code else lbl

label_freq = {}
for name in product_order:
    lbl = base_label(name)
    label_freq[lbl] = label_freq.get(lbl, 0) + 1
product_label = {}
for name in product_order:
    lbl = base_label(name)
    if label_freq[lbl] > 1:
        prefix = strip_noise_words(name.split(" (")[0].strip())
        lbl = f"{lbl} ({prefix})"
    product_label[name] = lbl

n_prod = len(product_order)

# ── ตารางแบบ pivot (ชีตรายบุคคล): วันที่เบิก + คอลัมน์แต่ละชนิดสายไฟ (จำนวน) + ค่าแรงรวมของวันนั้น ──
# แถวรวมท้ายตารางใช้สูตร =SUM(...) อ้างอิงแถวข้อมูลจริง ไม่ใช่ตัวเลขคงที่ — แก้ตัวเลขในแถวไหนใน Excel
# แล้วยอดรวม/ค่าแรงสุทธิท้ายชีตจะคำนวณตามให้อัตโนมัติ
def write_pivot_table(ws, row, rows_list):
    headers = ["วันที่เบิก"] + [product_label[n] for n in product_order] + ["ค่าแรง (บาท)"]
    widths = [17] + [11] * n_prod + [15]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        col = get_column_letter(ci)
        ws.column_dimensions[col].width = w
        is_prod_col = 2 <= ci <= 1 + n_prod
        if is_prod_col:
            pname = product_order[ci - 2]
            hexc = hexcolor(distinct_products[pname]) or "9CA3AF"
            fill, txt = hexc, contrast_text(hexc)
        else:
            fill, txt = NAVY, "FFFFFF"
        # wrap_text กันหัวคอลัมน์ยาวล้นออกไปทับคอลัมน์ข้างๆ ตอนคอลัมน์แคบลง (ตัดขึ้นบรรทัดใหม่แทน)
        cell(ws, f"{col}{row}", h, font=Font(name=FONT, size=FS(9.5), bold=True, color=txt), fill=fill, align=CW, border=box)
    ws.row_dimensions[row].height = RH(26)
    row += 1

    # คอลัมน์ซ่อนไว้ทางขวาของตาราง เก็บค่าแรงแยกตามชนิด x วันที่ — เป็นแหล่งอ้างอิงของสูตร SUM แนวตั้ง
    # ของแถว "ค่าแรงตัด (บาท)" ด้านล่าง (ตัวเลขค่าแรงต่อวันในคอลัมน์ที่มองเห็นเป็นยอดรวมทุกชนิดของวันนั้น
    # จึงต้องมีคอลัมน์แยกตามชนิดไว้ต่างหากเพื่อให้ SUM ตามชนิดได้ถูกต้อง)
    hidden_cols = {n: get_column_letter(LAST_P + 1 + i) for i, n in enumerate(product_order)}
    for col in hidden_cols.values():
        ws.column_dimensions[col].hidden = True

    # อัตราค่าแรง/หน่วยของแต่ละชนิด (คงที่ต่อสินค้า) — ใช้ผูกเป็นสูตร qty*rate ในคอลัมน์ที่ซ่อนไว้
    rates = {}
    date_agg = {}
    for r in rows_list:
        dt = r["issued_at"]
        pname = r["product_name"]
        e = date_agg.setdefault(dt, {"qty": {}, "wage": 0.0, "wage_by_prod": {}})
        e["qty"][pname] = e["qty"].get(pname, 0) + r["good_qty"]
        e["wage"] += r["wage"]
        e["wage_by_prod"][pname] = e["wage_by_prod"].get(pname, 0) + r["wage"]
        if pname not in rates:
            rates[pname] = r.get("wage_per_unit") or 0

    first_data_row = row
    for dt in sorted(date_agg.keys()):
        e = date_agg[dt]
        vals = [date_th(dt)] + [e["qty"].get(n, 0) for n in product_order] + [e["wage"]]
        for ci, v in enumerate(vals, start=1):
            col = get_column_letter(ci)
            is_prod_col = 2 <= ci <= 1 + n_prod
            is_wage_col = ci == 2 + n_prod
            fmt = NUM_Z if is_prod_col else (MONEY if is_wage_col else None)
            cell(ws, f"{col}{row}", v, font=Font(name=FONT, size=FS(9.5), color="111827"),
                 align=(R if (is_prod_col or is_wage_col) else L), border=box, fmt=fmt)
        # สูตร = จำนวน(อ้างอิงช่องที่มองเห็น) x อัตราค่าแรง/หน่วย — แก้จำนวนในตารางแล้วค่าแรงเปลี่ยนตามจริง
        # ส่วนต่างเล็กน้อยจากงานเสีย/หาย (ซึ่งไม่ได้แสดงแยกในตารางนี้) บวกเพิ่มเป็นค่าคงที่ต่อท้าย เพื่อให้ยอดรวมยังตรงเป๊ะ
        for ci, n in enumerate(product_order, start=2):
            qcol = get_column_letter(ci)
            hcol = hidden_cols[n]
            qty = e["qty"].get(n, 0)
            wage = e["wage_by_prod"].get(n, 0)
            rate = rates.get(n, 0)
            adj = wage - qty * rate
            formula = f"={qcol}{row}*{rate:g}" if abs(adj) < 0.005 else f"={qcol}{row}*{rate:g}+{adj:.2f}"
            cell(ws, f"{hcol}{row}", formula, fmt=MONEY)
        ws.row_dimensions[row].height = RH(16)
        row += 1
    last_data_row = row - 1

    # ── บรรทัดรวม (subtotal) ต่อคอลัมน์ — สูตร SUM อ้างอิงแถวข้อมูลด้านบน ──
    # ช่องค่าแรงรวมท้ายแถวนี้ไม่ต้องใส่ตัวเลขซ้ำ (ปล่อยว่างไว้) เพราะแถว "ค่าแรงตัด" ถัดไปมีสรุปยอดเดียวกันอยู่แล้ว
    col_totals = {n: 0 for n in product_order}
    wage_total = 0.0
    cell(ws, f"A{row}", "รวม", font=Font(name=FONT, size=FS(9.5), bold=True), align=R, border=box)
    if last_data_row >= first_data_row:
        for ci, n in enumerate(product_order, start=2):
            col = get_column_letter(ci)
            col_totals[n] = sum(date_agg[dt]["qty"].get(n, 0) for dt in date_agg)
            cell(ws, f"{col}{row}", f"=SUM({col}{first_data_row}:{col}{last_data_row})",
                 font=Font(name=FONT, size=FS(9.5), bold=True), align=R, border=box, fmt=NUM_Z)
        wage_total = sum(e["wage"] for e in date_agg.values())
    cell(ws, f"{LAST_P_LETTER}{row}", None, font=Font(name=FONT, size=FS(9.5), bold=True), align=R, border=box)
    ws.row_dimensions[row].height = RH(18)
    row += 1

    # ── แถบสีเขียวอ่อน: ค่าแรงตัด (บาท) แยกตามชนิด — เป็นแหล่งอ้างอิงยอดค่าแรงรวมเพียงจุดเดียว (ไม่ซ้ำกับแถว "รวม" ด้านบน) ──
    # แนวตั้ง: แต่ละช่อง = SUM คอลัมน์ที่ซ่อนไว้ของชนิดนั้น (first_data_row:last_data_row)
    # แนวนอน: ช่องรวมท้ายแถว = SUM ของทุกช่องชนิดในแถวนี้เอง
    LIGHT_GREEN = "DCFCE7"
    cell(ws, f"A{row}", "ค่าแรงตัด (บาท)", font=Font(name=FONT, size=FS(9.5), bold=True, color=GREEN), fill=LIGHT_GREEN, align=R, border=box)
    if last_data_row >= first_data_row:
        for ci, n in enumerate(product_order, start=2):
            col = get_column_letter(ci)
            hcol = hidden_cols[n]
            cell(ws, f"{col}{row}", f"=SUM({hcol}{first_data_row}:{hcol}{last_data_row})",
                 font=Font(name=FONT, size=FS(9.5), bold=True, color=GREEN), fill=LIGHT_GREEN, align=R, border=box, fmt=MONEY_Z)
        cell(ws, f"{LAST_P_LETTER}{row}", f"=SUM(B{row}:{LABEL_END_LETTER}{row})",
             font=Font(name=FONT, size=FS(9.5), bold=True, color=GREEN), fill=LIGHT_GREEN, align=R, border=box, fmt=MONEY)
    else:
        cell(ws, f"{LAST_P_LETTER}{row}", 0, font=Font(name=FONT, size=FS(9.5), bold=True, color=GREEN), fill=LIGHT_GREEN, align=R, border=box, fmt=MONEY)
    wage_total_ref = f"{LAST_P_LETTER}{row}"
    ws.row_dimensions[row].height = RH(18)
    row += 1

    return row, col_totals, wage_total, wage_total_ref

LAST_P = n_prod + 2  # วันที่เบิก + สินค้าแต่ละชนิด + ค่าแรง
LAST_P_LETTER = get_column_letter(LAST_P)
LABEL_END_LETTER = get_column_letter(LAST_P - 1)

# ── ชีตสรุปรวม (หน้าแรก) — รหัส/ชื่อ + แยกยอดค่าแรงตามชนิดสายไฟ + รวมสุทธิ ──
ws0 = wb.create_sheet(safe_sheet_name("สรุปรวม", used_names))
ws0.sheet_view.showGridLines = False
n_prod = len(product_order)
FIXED_COLS0 = 3  # รหัส, ชื่อ-สกุล, ชื่อเล่น
last_col = FIXED_COLS0 + n_prod + 1  # + สินค้าแต่ละชนิด + ค่าแรงสุทธิ
last_col_letter = get_column_letter(last_col)
ws0.merge_cells(f"A1:{last_col_letter}1")
cell(ws0, "A1", d.get("org_name", ""), font=Font(name=FONT, size=FS(13), bold=True, color=NAVY), align=CW)
ws0.merge_cells(f"A2:{last_col_letter}2")
cell(ws0, "A2", f"สรุปรายงานเบิกงาน/ส่งงาน — รอบจ่ายค่าแรงเดือน {month_th(d['month'])}", font=Font(name=FONT, size=FS(11), color=GREY), align=CW)
ws0.merge_cells(f"A3:{last_col_letter}3")
cell(ws0, "A3", f"เส้นตัดยอด (cut-off): {date_th(d['cutoff'])}  ·  งานที่คืนหลังจากนี้ยกไปจ่ายรอบเดือน {month_th(d['next_month'])}", font=Font(name=FONT, size=FS(9.5), italic=True, color=GREY), align=CW)
ws0.row_dimensions[1].height = RH(30)
ws0.row_dimensions[2].height = RH(26)
ws0.row_dimensions[3].height = RH(24)

hdr_row = 5
headers0 = ["รหัส", "ชื่อ-สกุล", "ชื่อเล่น"] + [product_label[n] for n in product_order] + ["ค่าแรงสุทธิรอบนี้ (บาท)"]
widths0 = [9, 26, 15] + [13] * n_prod + [22]
for ci, (h, w) in enumerate(zip(headers0, widths0), start=1):
    col = get_column_letter(ci)
    ws0.column_dimensions[col].width = w
    is_prod_col = FIXED_COLS0 + 1 <= ci <= FIXED_COLS0 + n_prod
    if is_prod_col:
        pname = product_order[ci - FIXED_COLS0 - 1]
        hexc = hexcolor(distinct_products[pname]) or "9CA3AF"
        fill, txt = hexc, contrast_text(hexc)
    else:
        fill, txt = NAVY, "FFFFFF"
    # wrap_text กันหัวคอลัมน์ยาวล้นออกไปทับคอลัมน์ข้างๆ หรือหลุดขอบหน้ากระดาษ
    cell(ws0, f"{col}{hdr_row}", h, font=Font(name=FONT, size=FS(9), bold=True, color=txt), fill=fill, align=CW, border=box)
ws0.row_dimensions[hdr_row].height = RH(28)

row = hdr_row + 1
first_member_row = row
grand_qty = {n: 0 for n in product_order}
for m in d["members"]:
    pw_map = {pw["name"]: pw for pw in m.get("product_wages", [])}
    vals = [m["member_code"], m["member_name"], m.get("member_nickname") or "-"]
    vals += [pw_map.get(n, {}).get("qty", 0) for n in product_order]
    vals += [m["total_wage"]]
    for ci, v in enumerate(vals, start=1):
        col = get_column_letter(ci)
        is_prod_col = FIXED_COLS0 + 1 <= ci <= FIXED_COLS0 + n_prod
        is_total_col = ci == last_col
        fmt = NUM_Z if is_prod_col else (MONEY_Z if is_total_col else None)
        cell(ws0, f"{col}{row}", v, font=Font(name=FONT, size=FS(9.5), color="111827"),
             align=(R if (is_prod_col or is_total_col) else L), border=box, fmt=fmt)
    for n in product_order:
        grand_qty[n] += pw_map.get(n, {}).get("qty", 0)
    ws0.row_dimensions[row].height = RH(17)
    row += 1
last_member_row = row - 1

# ── บรรทัดรวม (subtotal) ต่อคอลัมน์ — สูตร SUM อ้างอิงแถวสมาชิกด้านบน ──
ws0.merge_cells(f"A{row}:C{row}")
cell(ws0, f"A{row}", "รวมทั้งหมด", font=Font(name=FONT, size=FS(10), bold=True, color="FFFFFF"), fill=GREEN, align=R, border=box)
has_members = last_member_row >= first_member_row
for ci, n in enumerate(product_order, start=FIXED_COLS0 + 1):
    col = get_column_letter(ci)
    v = f"=SUM({col}{first_member_row}:{col}{last_member_row})" if has_members else 0
    cell(ws0, f"{col}{row}", v, font=Font(name=FONT, size=FS(10), bold=True, color="FFFFFF"), fill=GREEN, align=R, fmt=NUM_Z, border=box)
wage_v = f"=SUM({last_col_letter}{first_member_row}:{last_col_letter}{last_member_row})" if has_members else 0
cell(ws0, f"{last_col_letter}{row}", wage_v, font=Font(name=FONT, size=FS(10), bold=True, color="FFFFFF"), fill=GREEN, align=R, fmt=MONEY, border=box)
ws0.row_dimensions[row].height = RH(20)

ws0.print_area = f"A1:{last_col_letter}{row}"
ws0.page_setup.paperSize = PAPER_SIZE_CODE[paper_size]
ws0.page_setup.orientation = "landscape"
ws0.page_setup.fitToWidth = 1
ws0.page_setup.fitToHeight = 0
ws0.sheet_properties.pageSetUpPr.fitToPage = True
ws0.print_options.horizontalCentered = True
ws0.page_margins.left = ws0.page_margins.right = 0.35

# ── ชีตรายบุคคล — พยายามอัดให้พอดี 1 หน้ากระดาษ/คน ──
CONFIRM_TEXT = "ข้าพเจ้าขอยืนยันว่ารายการและจำนวนเงินดังกล่าวข้างต้นมีความถูกต้องครบถ้วนทุกประการ และได้รับเงินเรียบร้อยแล้ว"

def write_member_sheet(m, label=None):
    sheet_label = f'{m["member_code"]} {m["member_name"]}' + (f' {label}' if label else '')
    ws = wb.create_sheet(safe_sheet_name(sheet_label, used_names))
    ws.sheet_view.showGridLines = False

    # หัวชีต — ถ้ามี label (ต้นฉบับ/คู่ฉบับ) กันคอลัมน์ขวาสุดไว้เป็นป้ายมุมขวาบน ให้เห็นชัดแยกจากหัวเรื่องหลัก
    ws.merge_cells(f"A1:{LAST_P_LETTER}1")
    cell(ws, "A1", d.get("org_name", ""), font=Font(name=FONT, size=FS(13), bold=True, color=NAVY), align=CW)
    ws.merge_cells(f"A2:{LAST_P_LETTER}2")
    subtitle = f"รายงานเบิกงาน/ส่งงานรายบุคคล — รอบจ่ายค่าแรงเดือน {month_th(d['month'])}"
    cell(ws, "A2", subtitle, font=Font(name=FONT, size=FS(11), color=GREY), align=CW)
    ws.merge_cells(f"A3:{LAST_P_LETTER}3")
    cell(ws, "A3", f"เส้นตัดยอด (cut-off): {date_th(d.get('cutoff_start'))} - {date_th(d['cutoff'])}",
         font=Font(name=FONT, size=FS(9.5), italic=True, color=GREY), align=CW)
    ws.row_dimensions[1].height = RH(30)
    ws.row_dimensions[2].height = RH(26)
    ws.row_dimensions[3].height = RH(18)

    ws.merge_cells(f"A4:{LAST_P_LETTER}4")
    cell(ws, "A4", f'{m["member_code"]}   {m["member_name"]}' + (f'  ({m["member_nickname"]})' if m.get("member_nickname") else ''),
         font=Font(name=FONT, size=FS(11), bold=True, color="111827"), align=LW)
    ws.row_dimensions[4].height = RH(20)
    ws.merge_cells(f"A5:{LAST_P_LETTER}5")
    cell(ws, "A5", f'ธนาคาร: {m.get("bank_name") or "-"}   เลขบัญชี: {m.get("bank_account") or "-"}',
         font=Font(name=FONT, size=FS(9.5), color=GREY), align=L)

    row = 7
    row, _col_totals, _wage_total, wage_total_ref = write_pivot_table(ws, row, m["rows"])

    net_formula_parts = [wage_total_ref]
    if m.get("ng_deduction"):
        ws.merge_cells(f"A{row}:{LABEL_END_LETTER}{row}")
        cell(ws, f"A{row}", f'หัก NG เกินเกณฑ์ ({m["ng_excess_qty"]:g} เส้น × {d.get("ng_penalty_rate", 20):g} บาท)',
             font=Font(name=FONT, size=FS(10), color=RED), align=R, border=box)
        # แสดงเป็นสูตรคูณตรงๆ (จำนวนเกิน x อัตราค่าปรับ) ให้เห็นที่มาของตัวเลข ไม่ใช่แค่ผลลัพธ์สำเร็จรูป
        ng_ref = f"{LAST_P_LETTER}{row}"
        cell(ws, ng_ref, f'=-({m["ng_excess_qty"]:g}*{d.get("ng_penalty_rate", 20):g})',
             font=Font(name=FONT, size=FS(10), color=RED), align=R, fmt=MONEY, border=box)
        net_formula_parts.append(ng_ref)
        row += 1

    ws.merge_cells(f"A{row}:{LABEL_END_LETTER}{row}")
    cell(ws, f"A{row}", "ค่าแรงสุทธิรอบนี้", font=Font(name=FONT, size=FS(11), bold=True, color="FFFFFF"), fill=GREEN, align=R, border=box)
    # ปัดขึ้นเต็มบาทเหมือนสูตรฝั่งระบบ (Math.ceil) — ใช้ ROUNDUP แทน (ค่าแรงเป็นบวกเสมอ ผลเหมือนกัน)
    net_formula = f"=ROUNDUP({'+'.join(net_formula_parts)},0)"
    cell(ws, f"{LAST_P_LETTER}{row}", net_formula, font=Font(name=FONT, size=FS(11), bold=True, color="FFFFFF"), fill=GREEN, align=R, fmt=MONEY, border=box)
    ws.row_dimensions[row].height = RH(20)
    row += 2

    if m.get("carry_rows"):
        ws.merge_cells(f"A{row}:{LAST_P_LETTER}{row}")
        cell(ws, f"A{row}", f'งานที่คืนหลังเส้นตัดยอด ({date_th(d["cutoff"])}) — ยกยอดไปจ่ายรอบเดือน {month_th(d["next_month"])}',
             font=Font(name=FONT, size=FS(10), bold=True, color="FFFFFF"), fill=AMBER, align=LW)
        ws.row_dimensions[row].height = RH(28)
        row += 1
        row, _carry_col_totals, _carry_wage_total, _carry_wage_ref = write_pivot_table(ws, row, m["carry_rows"])

    # ── ช่องเซ็นรับเงิน ──
    row += 2
    ws.merge_cells(f"A{row}:{LAST_P_LETTER}{row}")
    cell(ws, f"A{row}", CONFIRM_TEXT, font=Font(name=FONT, size=FS(10), italic=True, color="111827"), align=RW)
    ws.row_dimensions[row].height = RH(18)
    row += 2
    ws.merge_cells(f"A{row}:{LAST_P_LETTER}{row}")
    cell(ws, f"A{row}", "ลงชื่อ .......................................................... ผู้รับเงิน",
         font=Font(name=FONT, size=FS(10.5)), align=R)
    row += 2
    ws.merge_cells(f"A{row}:{LAST_P_LETTER}{row}")
    cell(ws, f"A{row}", "วันที่ ............ / ............ / ............",
         font=Font(name=FONT, size=FS(10.5)), align=R)
    row += 1

    ws.print_area = f"A1:{LAST_P_LETTER}{row}"
    ws.page_setup.paperSize = PAPER_SIZE_CODE[paper_size]
    if paper_size == "A5":
        ws.page_setup.orientation = "landscape"
    else:
        ws.page_setup.orientation = "portrait" if n_prod <= 4 else "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top = ws.page_margins.bottom = 0.4

for m in d["members"]:
    if duplicate_for_pdf:
        write_member_sheet(m, label="ต้นฉบับ")
        if include_copy:
            write_member_sheet(m, label="คู่ฉบับ")
    else:
        write_member_sheet(m)

wb.save(out)
print(out)
