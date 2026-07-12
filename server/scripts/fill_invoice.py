# -*- coding: utf-8 -*-
# เติมข้อมูลลง template ใบแจ้งหนี้ (server/templates/invoice-template.xlsx)
# Usage: python fill_invoice.py <template.xlsx> <data.json> <out.xlsx>
import sys, json, datetime, warnings
warnings.simplefilter("ignore")
from openpyxl import load_workbook

tpl, dataf, out = sys.argv[1], sys.argv[2], sys.argv[3]
mode = sys.argv[4] if len(sys.argv) > 4 else ""       # "pdf" = เหลือเฉพาะชีตเอกสาร + ขยายฟอนต์
doctype = sys.argv[5] if len(sys.argv) > 5 else ""    # "receipt" = ทำ 2 หน้า (ต้นฉบับ/คู่ฉบับ)
d = json.load(open(dataf, encoding="utf-8-sig"))

wb = load_workbook(tpl)
ws = wb["ใบแจ้งหนี้"]

# ---- หัวบิล ----
if d.get("invoice_no"):
    try:    ws["J5"] = int(d["invoice_no"])
    except: ws["J5"] = d["invoice_no"]
dt = d.get("date")
if dt:
    _dd = None
    try:    _dd = datetime.datetime.strptime(dt[:10], "%Y-%m-%d")
    except: _dd = None
    ws["J6"] = _dd if _dd else dt
    # วันที่ลงชื่อ "ผู้รับเงิน" (J41) เดิมเป็น =TODAY() -> ตั้งให้ตรงกับวันที่หัวบิล
    if _dd is not None:
        ws["J41"] = _dd

# ป้าย "วันครบกำหนด / Due Date" (I9,I10) ตั้ง wrap ไว้แล้วถูกตัด -> หดให้พอดีช่อง เห็นครบ
from openpyxl.styles import Alignment as _Al
for _c in ("I9", "I10"):
    _a = ws[_c].alignment
    ws[_c].alignment = _Al(horizontal=_a.horizontal, vertical=_a.vertical, wrap_text=False, shrink_to_fit=True)

cust = d.get("customer", {})
if cust.get("name"):    ws["D11"] = cust["name"]
if cust.get("address"): ws["D12"] = cust["address"]
if cust.get("contact"): ws["D13"] = cust["contact"]
if cust.get("taxid"):   ws["E14"] = cust["taxid"]

# ---- รายการสินค้า (แถว 18-29 = 12 บรรทัด) ----
BASE = 18
CAP = 12
lines = d.get("lines", [])[:CAP]
for i in range(CAP):
    r = BASE + i
    if i < len(lines):
        l = lines[i]
        ws.cell(row=r, column=3).value = l.get("project") or None        # C โครงการ
        ws.cell(row=r, column=4).value = l.get("part_number") or None    # D รหัสสินค้า
        ws.cell(row=r, column=5).value = l.get("description") or None    # E รายการ (merged E:G)
        ws.cell(row=r, column=8).value = l.get("quantity") or None       # H จำนวนหน่วย
        ws.cell(row=r, column=9).value = l.get("price") or None          # I ราคาต่อหน่วย
        # B (ลำดับ) และ J (จำนวนเงิน) ปล่อยเป็นสูตรเดิมของ template ให้คำนวณเอง
    else:
        # แถวว่าง: ล้างสูตร lookup เดิมออก ไม่ให้โชว์ค่าค้าง
        for col in (3, 4, 5, 8, 9):
            ws.cell(row=r, column=col).value = None

# โหมด PDF: ลบชีตข้อมูลช่วยออก เหลือเฉพาะเอกสาร + ขยายฟอนต์ ~15% (fit กว้าง 1 หน้าเหมือนเดิม)
if mode == "pdf":
    for sn in list(wb.sheetnames):
        if sn != "ใบแจ้งหนี้":
            del wb[sn]
    from openpyxl.worksheet.properties import PageSetupProperties
    from openpyxl.styles import Font, Border, Side

    def closeBox(sheet):
        # วาดขอบล่างของกล่อง (medium) ที่แถวสุดท้ายของ print_area (44) กันขอบล่างหาย
        med = Side(style='medium')
        for col in 'ABCDEFGHIJK':
            c = sheet[f'{col}44']; b = c.border
            c.border = Border(left=b.left, right=b.right, top=b.top, bottom=med)

    def scaleFonts(sheet):
        for row in sheet.iter_rows(min_row=1, max_row=55, min_col=1, max_col=14):
            for cell in row:
                f = cell.font
                cell.font = Font(name=f.name, size=round((f.size or 11) * 1.15, 1),
                                 bold=f.bold, italic=f.italic, color=f.color,
                                 underline=f.underline, strike=f.strike, vertAlign=f.vertAlign)

    def pageFit(sheet):
        if sheet.sheet_properties.pageSetUpPr is None:
            sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        else:
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1   # บังคับ 1 หน้า/แผ่น (ใบเสร็จ = 2 แผ่น ต้นฉบับ/คู่ฉบับ)
        sheet.page_setup.scale = None
        sheet.print_area = "A1:K44"        # ตัดแถวว่างท้ายออก → fit ได้ตัวใหญ่ขึ้น

    scaleFonts(ws)
    pageFit(ws)
    closeBox(ws)

    # ใบเสร็จรับเงิน: ทำ 2 หน้า — ต้นฉบับ (Original) + คู่ฉบับ (Copy)
    if doctype == "receipt":
        ws["J3"] = "ใบเสร็จรับเงิน (ต้นฉบับ)"
        ws["J4"] = "RECEIPT (Original)"
        ws2 = wb.copy_worksheet(ws)
        ws2.title = "copy"
        ws2["J3"] = "ใบเสร็จรับเงิน (คู่ฉบับ)"
        ws2["J4"] = "RECEIPT (Copy)"
        pageFit(ws2)
        closeBox(ws2)

wb.save(out)
print(out)
